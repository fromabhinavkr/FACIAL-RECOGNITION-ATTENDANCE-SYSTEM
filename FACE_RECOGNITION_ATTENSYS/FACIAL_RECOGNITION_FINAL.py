import face_recognition
import cv2
import numpy as np
import os
import openpyxl
from openpyxl import Workbook
import datetime
from datetime import date
import winsound
import pyttsx3
import time

def create_attendance_sheet(workbook, sheet_name, known_face_names):
    if worksheet_exists(workbook, sheet_name):
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value='Name')
        ws.cell(row=1, column=2, value=str(date.today()))
        ws.cell(row=1, column=3, value='Time')
        ws.column_dimensions['A'].width = 20 

        # To write names in a fixed order
        for row_num, name in enumerate(known_face_names, start=2):
            ws.cell(row=row_num, column=1, value=name)

    return ws

def is_session_ended():
    global current_session_end_time
    if current_session_end_time is not None:
        current_time = datetime.datetime.now()
        if current_time >= current_session_end_time:
            return True
    return False

def worksheet_exists(workbook, worksheet_name):
    try:
        workbook[worksheet_name]
        return True
    except KeyError:
        return False
engine = pyttsx3.init()
voices = engine.getProperty('voices')      
engine.setProperty('voice', voices[0].id)
engine.setProperty('rate', 150)

def start_session(session_duration):
    global current_session_start_time, current_session_end_time
    current_session_start_time = datetime.datetime.now()
    current_session_end_time = current_session_start_time + datetime.timedelta(minutes=session_duration)
    warning_given = False

def end_session():
    global current_session_start_time, current_session_end_time, attendees
    for name, times in attendees.items():
        total_duration = sum((end - start).total_seconds() for start, end in times)
        attendance_threshold = 0.75 * (current_session_end_time - current_session_start_time).total_seconds()
        if total_duration >= attendance_threshold:
            mark_attendance(name, "Present")
    current_session_start_time = None
    current_session_end_time = None
    attendees = {}
    warning_given = False
    return

def mark_attendance(name, status):
    global row
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=status)
    ws.cell(row=row, column=3, value=str(date.today()))
    row += 1
    wb.save('attendance_excel.xlsx')

#DATABASE MAKING###########################################################
CurrentFolder = os.getcwd() 
image = os.path.join(CurrentFolder, 'abhinav.jpg')
image2 = os.path.join(CurrentFolder, 'nafjadjaffer.jpeg')
image3 = os.path.join(CurrentFolder, 'ProfRameez.jpg')
image4 = os.path.join(CurrentFolder, 'hashir.jpg')
image5 = os.path.join(CurrentFolder, 'jeesonbrowny.png')
image6 = os.path.join(CurrentFolder, 'manudevmv.png')
image7 = os.path.join(CurrentFolder, 'sreeragk.png')
image8 = os.path.join(CurrentFolder, 'rahult.png')
image9 = os.path.join(CurrentFolder, 'deepakdpai.png')
image10 = os.path.join(CurrentFolder, 'sanjuantony.png')
image11 = os.path.join(CurrentFolder, 'ajnastp.png')
image12 = os.path.join(CurrentFolder, 'muzammil.jpg')
image13 = os.path.join(CurrentFolder, 'dinesh.jpg')
image14 = os.path.join(CurrentFolder, 'nufaiz.jpg')

video_capture = cv2.VideoCapture(0)

person1_name = "Abhinav KR"
person1_image = face_recognition.load_image_file(image)
person1_face_encoding = face_recognition.face_encodings(person1_image)[0]

person2_name = "Nafjad Jaffer"
person2_image = face_recognition.load_image_file(image2)
person2_face_encoding = face_recognition.face_encodings(person2_image)[0]

person3_name = "Professor Rameez A"
person3_image = face_recognition.load_image_file(image3)
person3_face_encoding = face_recognition.face_encodings(person3_image)[0]

person4_name = "Hashir N"
person4_image = face_recognition.load_image_file(image4)
person4_face_encoding = face_recognition.face_encodings(person4_image)[0]

person5_name = "Jeeson Browny"
person5_image = face_recognition.load_image_file(image5)
person5_face_encoding = face_recognition.face_encodings(person5_image)[0]

person6_name = "Manudev M V"
person6_image = face_recognition.load_image_file(image6)
person6_face_encoding = face_recognition.face_encodings(person6_image)[0]

person7_name = "Sreerag K"
person7_image = face_recognition.load_image_file(image7)
person7_face_encoding = face_recognition.face_encodings(person7_image)[0]

person8_name = "Rahul T"
person8_image = face_recognition.load_image_file(image8)
person8_face_encoding = face_recognition.face_encodings(person8_image)[0]

person9_name = "Deepak D Pai"
person9_image = face_recognition.load_image_file(image9)
person9_face_encoding = face_recognition.face_encodings(person9_image)[0]

person10_name = "Sanju Antony"
person10_image = face_recognition.load_image_file(image10)
person10_face_encoding = face_recognition.face_encodings(person10_image)[0]

person11_name = "Ajnas T P"
person11_image = face_recognition.load_image_file(image11)
person11_face_encoding = face_recognition.face_encodings(person11_image)[0]

person12_name = "Muzammil"
person12_image = face_recognition.load_image_file(image12)
person12_face_encoding = face_recognition.face_encodings(person12_image)[0]

person13_name = "Dinesh"
person13_image = face_recognition.load_image_file(image13)
person13_face_encoding = face_recognition.face_encodings(person13_image)[0]

person14_name = "Nufaiz N"
person14_image = face_recognition.load_image_file(image14)
person14_face_encoding = face_recognition.face_encodings(person14_image)[0]

known_face_encodings = [
    person1_face_encoding,
    person2_face_encoding,
    person3_face_encoding,
    person4_face_encoding,
    person5_face_encoding,
    person6_face_encoding,
    person7_face_encoding,
    person8_face_encoding,
    person9_face_encoding,
    person10_face_encoding,
    person11_face_encoding,
    person12_face_encoding,
    person13_face_encoding,
    person14_face_encoding
]
known_face_names = [
    person1_name,
    person2_name,
    person3_name,
    person4_name,
    person5_name,
    person6_name,
    person7_name,
    person8_name,
    person9_name,
    person10_name,
    person11_name,
    person12_name,
    person13_name,
    person14_name
]
############################################################################

names_entered = set()
current_time = datetime.datetime.now()
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True
current_session_start_time = None
current_session_end_time = None
warning_given = False
attendees = {}  
faces_detected = set()
wb = openpyxl.load_workbook('attendance_excel.xlsx', keep_vba=True)
inp = input('Please give current subject lecture name') + str(date.today())
ws = create_attendance_sheet(wb, inp, known_face_names)

if worksheet_exists(wb, inp):
    ws = create_attendance_sheet(wb, inp, known_face_names)
    existing_names = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
else:
    ws = wb.create_sheet(inp)
    existing_names = set()
    ws.cell(row=1, column=1, value='Name/Date')
    ws.cell(row=1, column=2, value='Presence')
    ws.cell(row=1, column=3, value=str(date.today()))  # newtime

row = ws.max_row + 1  
names_marked_current_session = set()

while True:
    ret, frame = video_capture.read()

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = small_frame[:, :, ::-1]

    if process_this_frame:
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []

        #A threshold is introduced to increase the accuracy of face detection ###################
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
            confidence = 0.0
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                confidence = 1.0 - face_distances[best_match_index]
            # Only accept the name if the confidence variable is above 0.6 threshold 
            if confidence > 0.6:
                face_names.append(name)
            else:
                face_names.append("Unknown")
        #########################################################################################

        for face_encoding, name in zip(face_encodings, face_names):
            if name != "Unknown":
                current_time = datetime.datetime.now() 
                if current_session_start_time is None:
                    # Start a new session if none is active
                    session_duration = int(input("Enter the session duration in minutes: "))
                    start_session(session_duration)

                if name not in attendees:
                    attendees[name] = []

                # Record entry time if not already present
                if not attendees[name] or attendees[name][-1][1] is None:
                    attendees[name].append((current_time, current_time))

                # Check if the session has ended
                if is_session_ended():
                    end_session()
                    if session_duration == 0:
                        break
                    # start_session(session_duration)
                if current_session_start_time is not None and current_session_end_time is not None:
                    if current_time >= current_session_start_time + (current_session_end_time - current_session_start_time) * 0.65 and not warning_given:
                        print("Attendance marking stops soon")
                        engine.say("Attendance marking stops soon")
                        engine.runAndWait()
                        warning_given = True

                row_num = ws.max_row + 1
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == name:
                        row_num = row
                        break

                if row_num <= ws.max_row:
                    current_time = datetime.datetime.now().strftime("%H:%M:%S")
                    ws.cell(row=row_num, column=2, value="Present")
                    ws.cell(row=row_num, column=3, value=current_time)
                    wb.save('attendance_excel.xlsx')
                    
                    
                    if name not in faces_detected:
                        faces_detected.add(name)
                        print(f"Attendance marked for {name}")
                        winsound.Beep(2500, 1000)
                        engine.say(name + ' present')
                        engine.runAndWait()
                    else:
                        print("Next student")
    process_this_frame = not process_this_frame


    # For getting a square box over the  faces with that detected face's name
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    cv2.imshow('Video', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):   
        end_session()
        print("data saved")
        break

video_capture.release()
cv2.destroyAllWindows()